from flask import Flask, jsonify, request
from collections import defaultdict
import os
import glob
import openpyxl
from flask_caching import Cache

app = Flask(__name__)

# Configuração de cache
cache = Cache(app, config={'CACHE_TYPE': 'simple'})

# Estrutura global para armazenar os dados processados
dados_ies = {}

def encontrar_arquivo_excel():
    """Encontra o arquivo Excel no diretório atual"""
    # Procurar por arquivos Excel com várias extensões possíveis
    extensoes = ['*.xlsx', '*.xls']
    for extensao in extensoes:
        arquivos = glob.glob(extensao)
        if arquivos:
            return arquivos[0]
    return None

def processar_arquivo_excel(nome_arquivo):
    """
    Processa o arquivo Excel e retorna um dicionário com os dados de todas as IES
    usando openpyxl em vez de pandas
    """
    try:
        print(f"Processando arquivo: {nome_arquivo}")
        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(nome_arquivo, data_only=True)
        ies_abas = wb.sheetnames
        
        print(f"Abas encontradas: {ies_abas}")
        dados_processados = {}
        
        for ies in ies_abas:
            print(f"Processando IES: {ies}")
            # Obter a aba
            sheet = wb[ies]
            
            # Encontrar os cabeçalhos
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value).strip() if cell_value is not None else f"Coluna{col}")
            
            print(f"Cabeçalhos encontrados: {headers}")
            
            # Verificar se as colunas necessárias existem
            colunas_necessarias = ['Semestre', 'Materia', 'Tema', 'Subtema', 'Aula']
            col_indices = {}
            
            for col_name in colunas_necessarias:
                found = False
                for idx, header in enumerate(headers):
                    if header.lower() == col_name.lower():
                        col_indices[col_name] = idx
                        found = True
                        break
                if not found:
                    print(f"Aviso: Coluna '{col_name}' não encontrada na IES {ies}")
                    col_indices[col_name] = -1
            
            # Estrutura hierárquica para os dados desta IES
            ies_estruturada = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
            
            # Processar cada linha (começando da linha 2)
            for row in range(2, sheet.max_row + 1):
                try:
                    # Obter valores das células
                    semestre_val = sheet.cell(row=row, column=col_indices['Semestre'] + 1).value
                    materia_val = sheet.cell(row=row, column=col_indices['Materia'] + 1).value
                    tema_val = sheet.cell(row=row, column=col_indices['Tema'] + 1).value
                    subtema_val = sheet.cell(row=row, column=col_indices['Subtema'] + 1).value
                    aula_val = sheet.cell(row=row, column=col_indices['Aula'] + 1).value
                    
                    semestre = str(semestre_val or "").strip()
                    materia = str(materia_val or "").strip()
                    tema = str(tema_val or "").strip()
                    subtema = str(subtema_val or "").strip()
                    aula = str(aula_val or "").strip()
                    
                    # Obter links se existirem
                    link_aula = ""
                    link_pdf = ""
                    link_quiz = ""
                    
                    # Procurar por colunas de links
                    for idx, header in enumerate(headers):
                        cell_val = sheet.cell(row=row, column=idx + 1).value
                        if cell_val is None:
                            continue
                            
                        header_lower = header.lower()
                        if "link" in header_lower and "aula" in header_lower:
                            link_aula = str(cell_val or "").strip()
                        elif "link" in header_lower and "pdf" in header_lower:
                            link_pdf = str(cell_val or "").strip()
                        elif "link" in header_lower and "quiz" in header_lower:
                            link_quiz = str(cell_val or "").strip()
                    
                    # Pular linhas com dados essenciais faltantes
                    if not all([semestre, materia, tema, subtema, aula]):
                        continue
                    
                    # Criar objeto de aula
                    aula_obj = {
                        'nome': aula,
                        'link_aula': link_aula if link_aula else None,
                        'link_pdf': link_pdf if link_pdf else None,
                        'link_quiz': link_quiz if link_quiz else None
                    }
                    
                    # Adicionar à estrutura hierárquica
                    ies_estruturada[semestre][materia][tema].append({
                        'subtema': subtema,
                        'aula': aula_obj
                    })
                    
                except Exception as e:
                    print(f"Erro ao processar linha {row} na IES {ies}: {e}")
                    continue
            
            # Converter defaultdict para dict regular e organizar a estrutura
            ies_estruturada_final = {}
            for semestre, materias in ies_estruturada.items():
                semestre_dict = {}
                for materia, temas in materias.items():
                    materia_dict = {}
                    for tema, subtemas in temas.items():
                        materia_dict[tema] = subtemas
                    semestre_dict[materia] = materia_dict
                ies_estruturada_final[semestre] = semestre_dict
            
            dados_processados[ies] = ies_estruturada_final
            
        return dados_processados
        
    except Exception as e:
        print(f"Erro ao processar arquivo Excel: {e}")
        import traceback
        traceback.print_exc()
        return {}

def formatar_resposta_api(dados_ies, especifica_ies=None, semestre=None):
    """
    Formata os dados para a resposta da API conforme a hierarquia solicitada
    """
    resultado = {}
    
    # Filtrar por IES específica se fornecida
    ies_para_processar = {especifica_ies: dados_ies[especifica_ies]} if especifica_ies and especifica_ies in dados_ies else dados_ies
    
    for ies_nome, ies_dados in ies_para_processar.items():
        ies_resultado = {}
        
        # Filtrar por semestre se fornecido
        if semestre and semestre in ies_dados:
            semestres_para_processar = {semestre: ies_dados[semestre]}
        else:
            semestres_para_processar = ies_dados
        
        for semestre_nome, semestre_dados in semestres_para_processar.items():
            semestre_resultado = []
            
            for materia, temas in semestre_dados.items():
                materia_dict = {
                    'materia': materia,
                    'temas': []
                }
                
                for tema, subtemas in temas.items():
                    tema_dict = {
                        'tema': tema,
                        'subtemas': []
                    }
                    
                    for subtema_info in subtemas:
                        subtema_dict = {
                            'subtema': subtema_info['subtema'],
                            'aulas': [subtema_info['aula']]
                        }
                        tema_dict['subtemas'].append(subtema_dict)
                    
                    materia_dict['temas'].append(tema_dict)
                
                semestre_resultado.append(materia_dict)
            
            ies_resultado[semestre_nome] = semestre_resultado
        
        resultado[ies_nome] = ies_resultado
    
    return resultado

# Endpoint raiz com informações da API
@app.route('/')
def home():
    # Gerar HTML com links clicáveis para todas as IES
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>API de Conteúdos de IES</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; }
            h1 { color: #333; }
            ul { list-style-type: none; padding: 0; }
            li { margin: 10px 0; }
            a { 
                text-decoration: none; 
                color: #0366d6; 
                font-weight: bold; 
                padding: 8px 12px;
                border: 1px solid #0366d6;
                border-radius: 4px;
                display: inline-block;
            }
            a:hover { background-color: #f0f7ff; }
            .endpoint { margin-top: 30px; }
            .ies-list { margin-top: 20px; }
            .warning { 
                background-color: #fff3cd; 
                border: 1px solid #ffeaa7; 
                color: #856404; 
                padding: 15px; 
                border-radius: 5px; 
                margin: 20px 0; 
            }
            .info {
                background-color: #d1ecf1; 
                border: 1px solid #bee5eb; 
                color: #0c5460; 
                padding: 15px; 
                border-radius: 5px; 
                margin: 20px 0; 
            }
        </style>
    </head>
    <body>
        <h1>API de Conteúdos de IES</h1>
        <p>API para acesso aos conteúdos das IES a partir de arquivo Excel.</p>
    """
    
    # Verificar se há dados carregados
    if not dados_ies:
        arquivo_excel = encontrar_arquivo_excel()
        
        if arquivo_excel:
            html += f"""
            <div class="info">
                <h2>Arquivo Excel encontrado: {arquivo_excel}</h2>
                <p>Clique no botão abaixo para carregar os dados:</p>
                <p><a href="/recarregar-dados" style="color: #fff; background-color: #28a745; padding: 10px 15px; border-radius: 4px;">Carregar Dados do Arquivo</a></p>
            </div>
            """
        else:
            html += """
            <div class="warning">
                <h2>Atenção: Nenhum arquivo Excel encontrado</h2>
                <p>Por favor, faça upload de um arquivo Excel (.xlsx ou .xls) com a estrutura correta no diretório do servidor.</p>
                <p>Diretório atual: """ + os.getcwd() + """</p>
            </div>
            """
    
    html += """
        <div class="endpoint">
            <h2>Endpoints disponíveis:</h2>
            <ul>
                <li><a href="/listar-ies">/listar-ies</a> - Lista todas as IES disponíveis</li>
                <li><code>/&lt;nome_ies&gt;</code> - Todos os conteúdos de uma IES</li>
                <li><code>/&lt;nome_ies&gt;/&lt;semestre&gt;</code> - Conteúdos de uma IES por semestre</li>
            </ul>
        </div>
    """
    
    # Adicionar lista de IES com links clicáveis se os dados foram carregados
    if dados_ies:
        html += """
        <div class="ies-list">
            <h2>IES Disponíveis (clique para acessar):</h2>
            <ul>
        """
        
        for ies in sorted(dados_ies.keys()):
            html += f'<li><a href="/{ies}">{ies}</a></li>'
        
        html += """
            </ul>
        </div>
        """
    
    html += """
    </body>
    </html>
    """
    
    return html

# Endpoint para listar todas as IES disponíveis
@app.route('/listar-ies')
@cache.cached(timeout=300)
def listar_ies():
    """Retorna a lista de todas as IES disponíveis na API"""
    if not dados_ies:
        return jsonify({"error": "Nenhum arquivo Excel carregado. Use /recarregar-dados para carregar os dados."}), 404
        
    ies_disponiveis = list(dados_ies.keys())
    return jsonify({"ies_disponiveis": ies_disponiveis})

# Rota dinâmica para acessar conteúdos por IES
@app.route('/<string:nome_ies>')
@cache.cached(timeout=300)
def get_conteudos_ies(nome_ies):
    """Retorna todos os conteúdos de uma IES específica"""
    if not dados_ies:
        return jsonify({"error": "Nenhum arquivo Excel carregado. Use /recarregar-dados para carregar os dados."}), 404
        
    if nome_ies not in dados_ies:
        return jsonify({"error": f"IES '{nome_ies}' não encontrada"}), 404
    
    dados_formatados = formatar_resposta_api(dados_ies, nome_ies)
    
    # Adicionar links para semestres se quisermos uma visualização HTML
    if request.args.get('format') == 'html':
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Conteúdos da IES {nome_ies}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                h1 {{ color: #333; }}
                ul {{ list-style-type: none; padding: 0; }}
                li {{ margin: 10px 0; }}
                a {{ 
                    text-decoration: none; 
                    color: #0366d6; 
                    font-weight: bold; 
                    padding: 8px 12px;
                    border: 1px solid #0366d6;
                    border-radius: 4px;
                    display: inline-block;
                }}
                a:hover {{ background-color: #f0f7ff; }}
                .back-link {{ margin-top: 20px; }}
            </style>
        </head>
        <body>
            <h1>Conteúdos da IES {nome_ies}</h1>
            <p><a href="/">← Voltar para página inicial</a></p>
            
            <h2>Semestres disponíveis:</h2>
            <ul>
        """
        
        for semestre in sorted(dados_ies[nome_ies].keys()):
            html += f'<li><a href="/{nome_ies}/{semestre}">Semestre {semestre}</a></li>'
        
        html += """
            </ul>
        </body>
        </html>
        """
        return html
    
    return jsonify(dados_formatados)

# Rota dinâmica para acessar conteúdos por IES e semestre
@app.route('/<string:nome_ies>/<string:semestre>')
@cache.cached(timeout=300)
def get_conteudos_ies_semestre(nome_ies, semestre):
    """Retorna os conteúdos de uma IES específica filtrados por semestre"""
    if not dados_ies:
        return jsonify({"error": "Nenhum arquivo Excel carregado. Use /recarregar-dados para carregar os dados."}), 404
        
    if nome_ies not in dados_ies:
        return jsonify({"error": f"IES '{nome_ies}' não encontrada"}), 404
    
    if semestre not in dados_ies[nome_ies]:
        return jsonify({"error": f"Semestre '{semestre}' não encontrado para a IES '{nome_ies}'"}), 404
    
    dados_formatados = formatar_resposta_api(dados_ies, nome_ies, semestre)
    
    # Adicionar visualização HTML se solicitado
    if request.args.get('format') == 'html':
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Conteúdos da IES {nome_ies} - Semestre {semestre}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                h1, h2 {{ color: #333; }}
                .back-link {{ margin-bottom: 20px; }}
                a {{ 
                    text-decoration: none; 
                    color: #0366d6; 
                }}
                a:hover {{ text-decoration: underline; }}
                .materia {{ margin-top: 20px; border-left: 4px solid #0366d6; padding-left: 15px; }}
                .tema {{ margin-left: 20px; }}
                .subtema {{ margin-left: 40px; }}
                .aula {{ margin-left: 60px; }}
            </style>
        </head>
        <body>
            <div class="back-link">
                <a href="/{nome_ies}">← Voltar para {nome_ies}</a> | 
                <a href="/">Página inicial</a>
            </div>
            
            <h1>IES {nome_ies} - Semestre {semestre}</h1>
        """
        
        for materia in dados_formatados[nome_ies][semestre]:
            html += f'<div class="materia"><h2>{materia["materia"]}</h2></div>'
            
            for tema in materia["temas"]:
                html += f'<div class="tema"><h3>{tema["tema"]}</h3></div>'
                
                for subtema in tema["subtemas"]:
                    html += f'<div class="subtema"><h4>{subtema["subtema"]}</h4></div>'
                    
                    for aula in subtema["aulas"]:
                        html += f'<div class="aula"><strong>{aula["nome"]}</strong>'
                        if aula["link_aula"]:
                            html += f' | <a href="{aula["link_aula"]}" target="_blank">Aula</a>'
                        if aula["link_pdf"]:
                            html += f' | <a href="{aula["link_pdf"]}" target="_blank">PDF</a>'
                        if aula["link_quiz"]:
                            html += f' | <a href="{aula["link_quiz"]}" target="_blank">Quiz</a>'
                        html += '</div>'
        
        html += """
        </body>
        </html>
        """
        return html
    
    return jsonify(dados_formatados)

# Endpoint para recarregar os dados sem reiniciar o servidor
@app.route('/recarregar-dados', methods=['POST', 'GET'])
def recarregar_dados():
    """Recarrega os dados do arquivo Excel sem precisar reiniciar o servidor"""
    global dados_ies
    try:
        arquivo_excel = encontrar_arquivo_excel()
        if not arquivo_excel:
            error_msg = "Nenhum arquivo .xlsx ou .xls encontrado no diretório"
            print(error_msg)
            
            if request.method == 'GET':
                return f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <title>Erro ao Recarregar</title>
                    <meta http-equiv="refresh" content="5;url=/" />
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 40px; }}
                        .error {{ 
                            background-color: #f8d7da; 
                            border: 1px solid #f5c6cb; 
                            color: #721c24; 
                            padding: 15px; 
                            border-radius: 5px; 
                        }}
                    </style>
                </head>
                <body>
                    <div class="error">
                        <h2>Erro ao recarregar dados</h2>
                        <p>{error_msg}</p>
                        <p>Diretório: {os.getcwd()}</p>
                        <p>Redirecionando para a página inicial em 5 segundos...</p>
                        <p><a href="/">Clique aqui se não for redirecionado</a></p>
                    </div>
                </body>
                </html>
                """, 404
                
            return jsonify({"error": error_msg}), 404
        
        dados_ies = processar_arquivo_excel(arquivo_excel)
        
        if not dados_ies:
            error_msg = "Erro ao processar o arquivo Excel. Verifique a estrutura do arquivo."
            print(error_msg)
            
            if request.method == 'GET':
                return f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <title>Erro ao Processar</title>
                    <meta http-equiv="refresh" content="5;url=/" />
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 40px; }}
                        .error {{ 
                            background-color: #f8d7da; 
                            border: 1px solid #f5c6cb; 
                            color: #721c24; 
                            padding: 15px; 
                            border-radius: 5px; 
                        }}
                    </style>
                </head>
                <body>
                    <div class="error">
                        <h2>Erro ao processar dados</h2>
                        <p>{error_msg}</p>
                        <p>Redirecionando para a página inicial em 5 segundos...</p>
                        <p><a href="/">Clique aqui se não for redirecionado</a></p>
                    </div>
                </body>
                </html>
                """, 500
                
            return jsonify({"error": error_msg}), 500
        
        # Se for uma requisição GET, redirecionar para a página inicial
        if request.method == 'GET':
            return f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Dados Recarregados</title>
                <meta http-equiv="refresh" content="3;url=/" />
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    .success {{ 
                        background-color: #d4edda; 
                        border: 1px solid #c3e6cb; 
                        color: #155724; 
                        padding: 15px; 
                        border-radius: 5px; 
                    }}
                </style>
            </head>
            <body>
                <div class="success">
                    <h2>Dados recarregados com sucesso!</h2>
                    <p>Arquivo: {arquivo_excel}</p>
                    <p>IES carregadas: {', '.join(list(dados_ies.keys()))}</p>
                    <p>Redirecionando para a página inicial em 3 segundos...</p>
                    <p><a href="/">Clique aqui se não for redirecionado</a></p>
                </div>
            </body>
            </html>
            """
        
        return jsonify({
            "status": "dados_recarregados", 
            "arquivo": arquivo_excel,
            "ies_carregadas": list(dados_ies.keys())
        })
    except Exception as e:
        error_msg = f"Erro ao recarregar dados: {str(e)}"
        print(error_msg)
        
        if request.method == 'GET':
            return f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Erro ao Recarregar</title>
                <meta http-equiv="refresh" content="5;url=/" />
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    .error {{ 
                        background-color: #f8d7da; 
                        border: 1px solid #f5c6cb; 
                        color: #721c24; 
                        padding: 15px; 
                        border-radius: 5px; 
                    }}
                </style>
                </head>
                <body>
                    <div class="error">
                        <h2>Erro ao recarregar dados</h2>
                        <p>{error_msg}</p>
                        <p>Redirecionando para a página inicial em 5 segundos...</p>
                        <p><a href="/">Clique aqui se não for redirecionado</a></p>
                    </div>
                </body>
            </html>
            """, 500
            
        return jsonify({"error": error_msg}), 500

# Handler para erros 404
@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": "Endpoint não encontrado"}), 404

# Handler para erros 500
@app.errorhandler(500)
def internal_error(error):
    return jsonify({"error": "Erro interno do servidor"}), 500

def carregar_dados_iniciais():
    """Carrega os dados iniciais ao iniciar o servidor"""
    global dados_ies
    arquivo_excel = encontrar_arquivo_excel()
    
    if arquivo_excel:
        print(f"Arquivo Excel encontrado: {arquivo_excel}")
        dados_ies = processar_arquivo_excel(arquivo_excel)
        