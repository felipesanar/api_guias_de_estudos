import openpyxl
from flask import Flask, jsonify, request
from collections import defaultdict
import os
import glob
from flask_caching import Cache

app = Flask(__name__)

# Configuração de cache
cache = Cache(app, config={'CACHE_TYPE': 'simple'})

# Estrutura global para armazenar os dados processados
dados_ies = {}

def processar_arquivo_excel(nome_arquivo):
    """
    Processa o arquivo Excel usando openpyxl (sem pandas)
    """
    try:
        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(nome_arquivo, data_only=True)
        ies_abas = wb.sheetnames
        
        dados_processados = {}
        
        for ies in ies_abas:
            ws = wb[ies]
            
            # Encontrar cabeçalhos
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"coluna_{col}")
            
            # Estrutura hierárquica para os dados desta IES
            ies_estruturada = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
            
            # Processar linhas (começando da linha 2)
            for row in range(2, ws.max_row + 1):
                try:
                    # Extrair valores das células
                    valores = {}
                    for col in range(1, ws.max_column + 1):
                        if col - 1 < len(headers):
                            cell_value = ws.cell(row=row, column=col).value
                            valores[headers[col - 1]] = str(cell_value).strip() if cell_value is not None else ""
                    
                    semestre = valores.get('Semestre', '').strip()
                    materia = valores.get('Materia', '').strip()
                    tema = valores.get('Tema', '').strip()
                    subtema = valores.get('Subtema', '').strip()
                    aula = valores.get('Aula', '').strip()
                    link_aula = valores.get('Link Aula', '').strip()
                    link_pdf = valores.get('Link PDF', '').strip()
                    link_quiz = valores.get('Link Quiz', '').strip()
                    
                    # Pular linhas com dados essenciais faltantes
                    if not all([semestre, materia, tema, subtema, aula]):
                        continue
                    
                    # Criar objeto de aula
                    aula_obj = {
                        'nome': aula,
                        'link_aula': link_aula if link_aula else None,
                        'link_pdf': link_pdf if link_pdf else None,
                        'link_quiz': link_quiz if link_quiz else None
                    }
                    
                    # Adicionar à estrutura hierárquica
                    ies_estruturada[semestre][materia][tema].append({
                        'subtema': subtema,
                        'aula': aula_obj
                    })
                    
                except Exception as e:
                    print(f"Erro ao processar linha {row} na IES {ies}: {e}")
                    continue
            
            # Converter defaultdict para dict regular e organizar a estrutura
            ies_estruturada_final = {}
            for semestre, materias in ies_estruturada.items():
                semestre_dict = {}
                for materia, temas in materias.items():
                    materia_dict = {}
                    for tema, subtemas in temas.items():
                        materia_dict[tema] = subtemas
                    semestre_dict[materia] = materia_dict
                ies_estruturada_final[semestre] = semestre_dict
            
            dados_processados[ies] = ies_estruturada_final
            
        return dados_processados
        
    except Exception as e:
        print(f"Erro ao processar arquivo Excel: {e}")
        return {}

# ... (o resto do código permanece EXATAMENTE igual - funções formatar_resposta_api, endpoints, etc.)

def formatar_resposta_api(dados_ies, especifica_ies=None, semestre=None):
    """
    Formata os dados para a resposta da API conforme a hierarquia solicitada
    """
    resultado = {}
    
    # Filtrar por IES específica se fornecida
    ies_para_processar = {especifica_ies: dados_ies[especifica_ies]} if especifica_ies and especifica_ies in dados_ies else dados_ies
    
    for ies_nome, ies_dados in ies_para_processar.items():
        ies_resultado = {}
        
        # Filtrar por semestre se fornecido
        if semestre and semestre in ies_dados:
            semestres_para_processar = {semestre: ies_dados[semestre]}
        else:
            semestres_para_processar = ies_dados
        
        for semestre_nome, semestre_dados in semestres_para_processar.items():
            semestre_resultado = []
            
            for materia, temas in semestre_dados.items():
                materia_dict = {
                    'materia': materia,
                    'temas': []
                }
                
                for tema, subtemas in temas.items():
                    tema_dict = {
                        'tema': tema,
                        'subtemas': []
                    }
                    
                    for subtema_info in subtemas:
                        subtema_dict = {
                            'subtema': subtema_info['subtema'],
                            'aulas': [subtema_info['aula']]
                        }
                        tema_dict['subtemas'].append(subtema_dict)
                    
                    materia_dict['temas'].append(tema_dict)
                
                semestre_resultado.append(materia_dict)
            
            ies_resultado[semestre_nome] = semestre_resultado
        
        resultado[ies_nome] = ies_resultado
    
    return resultado

# ... (mantenha TODOS os endpoints Flask exatamente como estão)

if __name__ == '__main__':
    # Procurar por arquivos Excel no diretório atual
    arquivos = glob.glob('*.xlsx')
    
    if not arquivos:
        print("Nenhum arquivo .xlsx encontrado no diretório atual.")
        print("Por favor, coloque um arquivo Excel com a estrutura especificada na mesma pasta do script.")
        # No Render, não queremos que a aplicação falhe completamente
        # Inicializamos com dados vazios mas a API fica disponível
        dados_ies = {}
    else:
        print(f"Processando arquivo: {arquivos[0]}")
        dados_ies = processar_arquivo_excel(arquivos[0])
        
        if dados_ies:
            print(f"Dados carregados com sucesso para as IES: {list(dados_ies.keys())}")
            print("API pronta para receber requisições.")
        else:
            print("Erro ao processar o arquivo Excel. Verifique a estrutura do arquivo.")
            dados_ies = {}
    
    # No Render, use a porta fornecida pela variável de ambiente PORT
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)